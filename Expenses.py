from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

def prompt_user():
    while True:
        user_input = input("Do you want to create a new Excel Workbook or load an Existing one? (Create/Load): ").lower().rstrip()
        if user_input == "create":
            create_workbook()
            break
        elif user_input == "load":
            path = input("Please Enter the Directory and Name of the Workbook to be Loaded(Enter Workbook Name if it is in the same directory as this script!): ") 
            load_work_book(path)
            break
        else:
            print("Invalid Input!")

def create_workbook():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expenses"
    sheet.append(["Date", "Time", "Expenses", "ExpenseDescription"])

    for col in range(1, 5):
        sheet.cell(1, col).font = Font(bold=True)

    wb.save("Expenses.xlsx")
    loc = "Expenses.xlsx"
    load_work_book(loc)

def load_work_book(path):
    wb = load_workbook(path)
    sheet = wb.active

    while True:
        expense = input("Please Enter the expense: ")
        des = input("Please enter the Description/Nature of the expense: ")
        
        date_time = datetime.now()
        date = date_time.strftime("%b-%d-%Y")
        time = date_time.strftime("%H:%M:%S")

        entry = [date, time, expense, des]
        sheet.append(entry)

        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(path)

        another = input("Would you like to Enter another Expense?(y/n) ").rstrip().lower()
        if another != "y": 
            print("ThankYou for using this Program!")
            exit()

def main():
    print("""
\t\t\t---WELCOME!---
This is a Program to write Expense Reports to Excel Workbooks.
Please make sure that you are using latest Version of MS Excel and that the file type is 'xlsx'.
Also make sure to close the File you will be Modifying in this Program to Avoid Errors!
""")
    prompt_user()

main()


